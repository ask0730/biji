import os
import sys
import time
import json
import re
import logging
from pathlib import Path
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.keys import Keys
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

if not SELENIUM_AVAILABLE:
    print("错误: 需要安装 selenium 库")
    print("请运行: pip install selenium")
    sys.exit(1)

if not PANDAS_AVAILABLE:
    print("错误: 需要安装 pandas 库")
    print("请运行: pip install pandas openpyxl")
    sys.exit(1)


def init_logger():
    """初始化日志"""
    log_dir = os.path.join(os.path.dirname(__file__), "logs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "doubao_table_extract.log")

    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(formatter)
    
    logger = logging.getLogger("doubao_extractor")
    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    return logger


logger = init_logger()


def safe_get_text(element, max_retries=3):
    """安全地获取元素文本，处理stale element错误"""
    for attempt in range(max_retries):
        try:
            return element.text
    except Exception as e:
            if "stale" in str(e).lower() and attempt < max_retries - 1:
                time.sleep(0.1)
                continue
            raise
    return ""


def safe_find_elements(driver, by, value, max_retries=3):
    """安全地查找元素，处理可能的错误"""
    for attempt in range(max_retries):
        try:
            return driver.find_elements(by, value)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(0.1)
                continue
            logger.warning(f"查找元素失败 ({by}={value}): {str(e)}")
            return []
    return []


def safe_get_all_text(driver, selectors, max_retries=3):
    """安全地获取所有文本内容，处理stale element错误"""
    all_text = ""
    for selector in selectors:
        for attempt in range(max_retries):
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    try:
                        if elem.is_displayed():
                            text = safe_get_text(elem)
                            if len(text) > len(all_text):
                                all_text = text
            except Exception as e:
                        if "stale" not in str(e).lower():
                            logger.warning(f"提取文本失败: {str(e)}")
                        continue
                break  # 成功获取，跳出重试循环
    except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(0.2)
                    continue
                if "stale" not in str(e).lower():
                    logger.warning(f"查找元素失败 ({selector}): {str(e)}")
                break
    return all_text


def setup_driver():
    """设置浏览器驱动"""
    try:
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.implicitly_wait(10)
        return driver
    except Exception as e:
        logger.error(f"浏览器启动失败: {e}")
        raise




def upload_file_to_doubao(driver, file_path, wait_after_upload=True):
    """上传文件到豆包"""
    try:
        print(f"正在上传文件: {os.path.basename(file_path)}")
        
        # 查找附件按钮
        attachment_button = None
        
        # 方法1: 通过输入框查找附件按钮
        try:
            input_area = driver.find_element(By.CSS_SELECTOR, "textarea, input[type='text'], [contenteditable='true']")
            parent = input_area.find_element(By.XPATH, "./..")
            buttons = parent.find_elements(By.TAG_NAME, "button")
            for btn in buttons:
                if btn.is_displayed():
                    try:
                        svg = btn.find_element(By.TAG_NAME, "svg")
                        svg_html = svg.get_attribute('outerHTML')
                        if 'paperclip' in svg_html.lower() or 'attach' in svg_html.lower():
                            attachment_button = btn
                            break
                    except:
                        pass
        except:
            pass
        
        # 方法2: 直接查找附件按钮
        if not attachment_button:
            selectors = [
                "button svg[class*='paperclip']",
                "button svg[class*='attach']",
                "button[aria-label*='附件']",
                "button[aria-label*='上传']",
            ]
            for selector in selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        if elem.tag_name == 'svg':
                            try:
                                parent_btn = elem.find_element(By.XPATH, "./ancestor::button")
                                if parent_btn.is_displayed():
                                    attachment_button = parent_btn
                                    break
                            except:
                                continue
                        elif elem.tag_name == 'button' and elem.is_displayed():
                            attachment_button = elem
                            break
                    if attachment_button:
                        break
                except:
                    continue
        
        if attachment_button:
            print("找到附件按钮，正在点击...")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", attachment_button)
            time.sleep(0.5)
            attachment_button.click()
            time.sleep(1.5)
        
        # 查找文件输入框
        wait = WebDriverWait(driver, 5)
        try:
            file_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
        except TimeoutException:
            file_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
            if file_inputs:
                file_input = file_inputs[-1]
            else:
                raise Exception("未找到文件输入框")
        
        file_input.send_keys(os.path.abspath(file_path))
        print("✓ 文件上传成功")
        if wait_after_upload:
        time.sleep(3)  # 等待文件上传完成
        
        return True
        
    except Exception as e:
        logger.error(f"上传文件失败: {str(e)}")
        raise


def send_message_to_doubao(driver, message):
    """向豆包发送消息"""
    try:
        # 查找输入框并输入消息
        print(f"正在输入消息: {message}")
        
        # 查找输入框
        input_selectors = [
            "textarea",
            "input[type='text']",
            "[contenteditable='true']",
            "[placeholder*='发消息']",
            "[placeholder*='消息']",
        ]
        
        input_element = None
        for selector in input_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    if elem.is_displayed() and elem.is_enabled():
                        input_element = elem
                        break
                if input_element:
                    break
            except:
                continue
        
        if input_element:
            # 清空输入框
            input_element.clear()
            time.sleep(0.5)
            
            # 输入消息
            if input_element.tag_name == 'textarea' or input_element.tag_name == 'input':
                input_element.send_keys(message)
            else:
                # contenteditable元素
                input_element.send_keys(message)
            
            print(f"✓ 已输入消息: {message}")
            time.sleep(1)
        else:
            print("⚠️  未找到输入框，尝试使用JavaScript输入...")
            # 使用JavaScript输入
            script = f"""
            var inputs = document.querySelectorAll('textarea, input[type="text"], [contenteditable="true"]');
            for (var i = 0; i < inputs.length; i++) {{
                if (inputs[i].offsetParent !== null) {{
                    inputs[i].focus();
                    inputs[i].value = '{message}';
                    inputs[i].dispatchEvent(new Event('input', {{ bubbles: true }}));
                    break;
                }}
            }}
            """
            driver.execute_script(script)
            time.sleep(1)
        
        # 查找并点击发送按钮
        print("正在查找并点击发送按钮...")
        send_selectors = [
            "button[aria-label*='发送']",
            "button[title*='发送']",
            "button[type='submit']",
            "button svg[class*='send']",
            "button:has(svg[class*='send'])",
        ]
        
        send_button = None
        for selector in send_selectors:
            try:
                send_buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                for btn in send_buttons:
                    if btn.is_displayed() and btn.is_enabled():
                        send_button = btn
                        break
                if send_button:
                    break
            except:
                continue
        
        # 如果找不到发送按钮，尝试通过回车键发送
        if send_button:
            print("找到发送按钮，正在点击...")
            send_button.click()
        else:
            print("未找到发送按钮，尝试按回车键发送...")
            if input_element:
                input_element.send_keys(Keys.RETURN)
            else:
                # 使用JavaScript模拟回车
                script = """
                var inputs = document.querySelectorAll('textarea, input[type="text"], [contenteditable="true"]');
                for (var i = 0; i < inputs.length; i++) {
                    if (inputs[i].offsetParent !== null) {
                        var event = new KeyboardEvent('keydown', {
                            key: 'Enter',
                            code: 'Enter',
                            keyCode: 13,
                            bubbles: true
                        });
                        inputs[i].dispatchEvent(event);
                        break;
                    }
                }
                """
                driver.execute_script(script)
        
        print("✓ 消息已发送")
        time.sleep(2)
        
        return True
        
    except Exception as e:
        logger.error(f"发送消息失败: {str(e)}")
        raise


def wait_for_response(driver, timeout=60):
    """等待豆包返回响应"""
    print("正在等待豆包处理并返回结果...")
    
    start_time = time.time()
    last_text_length = 0
    stable_count = 0
    
    while time.time() - start_time < timeout:
        try:
            # 查找最新的消息内容
            message_selectors = [
                "[class*='message']",
                "[class*='response']",
                "[class*='content']",
                "article",
                "[role='article']",
            ]
            
            # 使用安全的方法获取文本
            all_text = safe_get_all_text(driver, message_selectors).strip()
            
            current_length = len(all_text)
            
            # 如果文本长度稳定（3秒内没有变化），认为响应完成
            if current_length == last_text_length and current_length > 0:
                stable_count += 1
                if stable_count >= 3:  # 3次检查都稳定
                    print("✓ 响应已完成")
                    return True
            else:
                stable_count = 0
                last_text_length = current_length
            
            time.sleep(1)
            
        except Exception as e:
            logger.warning(f"等待响应时出错: {str(e)}")
            time.sleep(1)
    
    print("⚠️  等待响应超时，将尝试提取已有内容")
    return False


def extract_structured_data_from_response(driver):
    """从响应中提取结构化数据（优先表格，因为豆包返回的是表格格式）"""
    try:
        # 优先提取表格数据（豆包返回的是表格格式）
        print("正在提取表格数据...")
        df = extract_table_from_response(driver)
        if df is not None and len(df) > 0:
            return df, 'table'
        
        # 如果表格提取失败，尝试提取JSON
        print("表格提取失败，尝试提取JSON数据...")
        json_data = extract_json_from_response(driver)
        if json_data:
            return json_data, 'json'
        
        return None, None
        
    except Exception as e:
        logger.error(f"提取结构化数据失败: {str(e)}")
        return None, None


def extract_table_from_response(driver):
    """从响应中提取表格数据"""
    try:
        print("正在提取表格数据...")
        
        # 方法1: 查找HTML表格（优先）
        try:
        tables = driver.find_elements(By.TAG_NAME, "table")
        except:
            tables = []
        
        if tables:
            print(f"找到 {len(tables)} 个HTML表格")
            all_dataframes = []
            
            # 提取所有表格
            for table_idx in range(len(tables)):
                try:
                    # 重新查找表格，避免stale element
                    current_tables = driver.find_elements(By.TAG_NAME, "table")
                    if table_idx >= len(current_tables):
                        continue
                    table = current_tables[table_idx]
                    
                    # 立即获取表格的HTML内容，避免后续stale element
                    try:
                        table_html = table.get_attribute('outerHTML')
                    except:
                        # 如果获取HTML失败，尝试直接提取
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                        except:
                            rows = None
                    else:
                        # 使用BeautifulSoup解析HTML（如果可用），否则回退到直接提取
                        try:
                            from bs4 import BeautifulSoup
                            soup = BeautifulSoup(table_html, 'html.parser')
                            rows_html = soup.find_all('tr')
                            rows = None  # 标记使用HTML解析
                        except:
                            rows = table.find_elements(By.TAG_NAME, "tr")
                    
                    if rows is None:
                        # 使用HTML解析
                        data = []
                        headers = []
                        for i, row_html in enumerate(rows_html):
                            cells_html = row_html.find_all(['td', 'th'])
                            if cells_html:
                                row_data = [cell.get_text(strip=True) for cell in cells_html]
                                if i == 0:
                                    if row_html.find('th'):
                                        headers = row_data
                                    else:
                                        headers = row_data
                                        data.append(row_data)
                                else:
                                    data.append(row_data)
                    else:
                        # 直接提取（原有方法，但立即获取文本）
                    if not rows:
                        continue
                    
                    data = []
                    headers = []
                    
                        for i in range(len(rows)):
                            try:
                                # 重新查找行，避免stale element
                                current_rows = table.find_elements(By.TAG_NAME, "tr")
                                if i >= len(current_rows):
                                    break
                                row = current_rows[i]
                                
                                # 立即获取单元格文本
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if not cells:
                            cells = row.find_elements(By.TAG_NAME, "th")
                        
                        if cells:
                                    # 立即获取所有单元格的文本（使用安全方法）
                                    row_data = []
                                    for cell in cells:
                                        try:
                                            text = safe_get_text(cell)
                                            row_data.append(text.strip())
                                        except:
                                            row_data.append('')
                                    
                            if i == 0:
                                        # 检查是否是表头
                                        try:
                                            has_th = len(row.find_elements(By.TAG_NAME, "th")) > 0
                                        except:
                                            has_th = False
                                        
                                        if has_th:
                                    headers = row_data
                                else:
                                    headers = row_data
                                    data.append(row_data)
                            else:
                                data.append(row_data)
                            except Exception as e:
                                logger.warning(f"提取第 {i+1} 行失败: {str(e)}")
                                continue
                    
                    if data or headers:
                        # 如果只有表头没有数据，使用表头作为第一行
                        if not data and headers:
                            data = [headers]
                            headers = []
                        
                        # 创建DataFrame
                        if headers:
                            # 确保列数一致
                            max_cols = max(len(headers), max([len(row) for row in data] if data else [0]))
                            headers = headers + [''] * (max_cols - len(headers))
                            for i, row in enumerate(data):
                                data[i] = row + [''] * (max_cols - len(row))
                            df = pd.DataFrame(data, columns=headers[:max_cols])
                        else:
                            # 没有表头，使用第一行作为表头
                            if data:
                                max_cols = max([len(row) for row in data])
                                for i, row in enumerate(data):
                                    data[i] = row + [''] * (max_cols - len(row))
                                df = pd.DataFrame(data[1:], columns=data[0][:max_cols] if data else None)
                            else:
                                continue
                        
                        if len(df) > 0:
                            all_dataframes.append(df)
                            print(f"  表格 {table_idx + 1}: {len(df)} 行 × {len(df.columns)} 列")
                except Exception as e:
                    logger.warning(f"提取表格 {table_idx + 1} 失败: {str(e)}")
                    continue
            
            # 如果有多个表格，合并它们
            if all_dataframes:
                if len(all_dataframes) == 1:
                    df = all_dataframes[0]
                else:
                    # 合并多个表格（垂直堆叠）
                    print(f"合并 {len(all_dataframes)} 个表格...")
                    df = pd.concat(all_dataframes, ignore_index=True)
                
                logger.info(f"从HTML表格提取了 {len(df)} 行数据")
                return df
        
        # 方法2: 从Markdown表格提取
        print("尝试从Markdown格式提取表格...")
        message_selectors = [
            "[class*='message']",
            "[class*='response']",
            "[class*='content']",
            "[class*='markdown']",
            "article",
            "[role='article']",
            "[class*='prose']",
        ]
        
        # 使用安全的方法获取文本
        all_text = safe_get_all_text(driver, message_selectors)
        
        # 从完整文本中提取所有表格
        if '|' in all_text:
            df = parse_markdown_table(all_text)
            if df is not None and len(df) > 0:
                print(f"从Markdown表格提取了 {len(df)} 行数据")
                logger.info(f"从Markdown表格提取了 {len(df)} 行数据")
                return df
        
        # 方法3: 从代码块中的表格提取
        print("尝试从代码块提取表格...")
        try:
            code_blocks = driver.find_elements(By.TAG_NAME, "pre")
            for code_block in code_blocks:
                try:
                if code_block.is_displayed():
                        # 立即获取文本
                    text = code_block.text
                    if '|' in text:
                        df = parse_markdown_table(text)
                        if df is not None and len(df) > 0:
                            print(f"从代码块表格提取了 {len(df)} 行数据")
                            logger.info(f"从代码块表格提取了 {len(df)} 行数据")
                            return df
                except Exception as e:
                    if "stale" in str(e).lower():
                        continue
                    logger.warning(f"提取代码块失败: {str(e)}")
                    continue
        except Exception as e:
            logger.warning(f"从代码块提取表格失败: {str(e)}")
        
        # 方法4: 尝试从所有文本中提取结构化数据
        print("尝试从文本内容提取结构化数据...")
        try:
            # 获取所有可见文本（使用安全方法）
            try:
            body = driver.find_element(By.TAG_NAME, "body")
                full_text = safe_get_text(body)
            except Exception as e:
                logger.warning(f"获取body文本失败: {str(e)}")
                full_text = ""
            
            # 如果文本中包含类似表格的结构（多行，每行有分隔符）
            lines = [line.strip() for line in full_text.split('\n') if line.strip()]
            if len(lines) > 2:
                # 检查是否有表格特征（多行，每行长度相似）
                potential_table_lines = []
                for line in lines:
                    # 检查是否包含多个字段（通过制表符、多个空格或特殊字符分隔）
                    if '\t' in line or '  ' in line or '|' in line:
                        potential_table_lines.append(line)
                
                if len(potential_table_lines) >= 2:
                    # 尝试解析为表格
                    data = []
                    for line in potential_table_lines:
                        # 尝试不同的分隔符
                        if '|' in line:
                            cells = [c.strip() for c in line.split('|') if c.strip()]
                        elif '\t' in line:
                            cells = line.split('\t')
                        else:
                            # 多个空格分隔
                            cells = [c.strip() for c in line.split('  ') if c.strip()]
                        
                        if len(cells) >= 2:  # 至少2列
                            data.append(cells)
                    
                    if data:
                        # 确保所有行列数一致
                        max_cols = max([len(row) for row in data])
                        for i, row in enumerate(data):
                            data[i] = row + [''] * (max_cols - len(row))
                        
                        df = pd.DataFrame(data[1:], columns=data[0][:max_cols] if data else None)
                        if len(df) > 0:
                            print(f"从文本结构提取了 {len(df)} 行数据")
                            logger.info(f"从文本结构提取了 {len(df)} 行数据")
                            return df
        except Exception as e:
            logger.warning(f"从文本提取表格失败: {str(e)}")
        
        print("⚠️  未找到表格数据，请检查豆包是否返回了表格格式的数据")
        return None
        
    except Exception as e:
        logger.error(f"提取表格失败: {str(e)}")
        return None


def parse_markdown_table(text):
    """解析Markdown格式的表格"""
    try:
        lines = text.split('\n')
        table_lines = []
        in_table = False
        
        for line in lines:
            stripped = line.strip()
            # 检测表格行（包含|分隔符）
            if '|' in stripped and not stripped.startswith('|---'):
                table_lines.append(stripped)
                in_table = True
            elif in_table and '|' not in stripped:
                break
        
        if not table_lines:
            return None
        
        # 解析表格
        rows = []
        for line in table_lines:
            # 移除首尾的|
            cells = [cell.strip() for cell in line.split('|')]
            # 过滤空字符串
            cells = [c for c in cells if c]
            if cells:
                rows.append(cells)
        
        if len(rows) < 2:
            return None
        
        # 第一行作为表头
        headers = rows[0]
        data = rows[1:]
        
        # 确保所有行的列数一致
        max_cols = max(len(headers), max([len(row) for row in data] if data else [0]))
        headers = headers + [''] * (max_cols - len(headers))
        
        for i, row in enumerate(data):
            data[i] = row + [''] * (max_cols - len(row))
        
        df = pd.DataFrame(data, columns=headers[:max_cols])
        return df
        
    except Exception as e:
        logger.error(f"解析Markdown表格失败: {str(e)}")
        return None


def extract_json_from_response(driver):
    """从响应中提取JSON格式的结构化数据"""
    try:
        print("正在提取JSON格式的结构化数据...")
        
        # 获取所有可见文本
        message_selectors = [
            "[class*='message']",
            "[class*='response']",
            "[class*='content']",
            "[class*='markdown']",
            "article",
            "[role='article']",
            "[class*='prose']",
            "pre",
            "code",
        ]
        
        # 使用安全的方法获取文本
        all_text = safe_get_all_text(driver, message_selectors)
        
        if not all_text:
            return None
        
        # 尝试提取JSON
        # 方法1: 查找代码块中的JSON
        json_pattern = r'```(?:json)?\s*(\[.*?\]|\{.*?\})\s*```'
        matches = re.findall(json_pattern, all_text, re.DOTALL)
        if matches:
            try:
                data = json.loads(matches[0])
                logger.info(f"从代码块提取了JSON数据")
                return data if isinstance(data, list) else [data]
            except:
                pass
        
        # 方法2: 查找纯JSON（在```之间或独立存在）
        json_patterns = [
            r'(\[[\s\S]*?\])',  # JSON数组
            r'(\{[\s\S]*?\})',  # JSON对象
        ]
        
        for pattern in json_patterns:
            matches = re.findall(pattern, all_text, re.DOTALL)
            for match in matches:
                try:
                    data = json.loads(match)
                    logger.info(f"从文本提取了JSON数据")
                    return data if isinstance(data, list) else [data]
                except:
                    continue
        
        # 方法3: 尝试从整个文本中提取JSON
        try:
            # 查找可能的JSON开始和结束位置
            start_idx = all_text.find('[')
            if start_idx == -1:
                start_idx = all_text.find('{')
            
            if start_idx != -1:
                # 尝试找到匹配的结束位置
                bracket_count = 0
                brace_count = 0
                in_string = False
                escape_next = False
                
                for i in range(start_idx, len(all_text)):
                    char = all_text[i]
                    
                    if escape_next:
                        escape_next = False
                        continue
                    
                    if char == '\\':
                        escape_next = True
                        continue
                    
                    if char == '"' and not escape_next:
                        in_string = not in_string
                        continue
                    
                    if not in_string:
                        if char == '[':
                            bracket_count += 1
                        elif char == ']':
                            bracket_count -= 1
                        elif char == '{':
                            brace_count += 1
                        elif char == '}':
                            brace_count -= 1
                        
                        if bracket_count == 0 and brace_count == 0 and i > start_idx:
                            json_str = all_text[start_idx:i+1]
                            try:
                                data = json.loads(json_str)
                                logger.info(f"从完整文本提取了JSON数据")
                                return data if isinstance(data, list) else [data]
                            except:
                                pass
        except Exception as e:
            logger.warning(f"提取JSON时出错: {str(e)}")
        
        print("⚠️  未找到有效的JSON数据")
        return None
        
    except Exception as e:
        logger.error(f"提取JSON失败: {str(e)}")
        return None


def map_to_template_format(data_list, template_path):
    """将提取的数据映射到模板格式"""
    try:
        print(f"正在加载模板文件: {os.path.basename(template_path)}")
        # 读取模板文件（根据文件扩展名选择引擎）
        file_ext = os.path.splitext(template_path)[1].lower()
        if file_ext == '.xls':
            template_df = pd.read_excel(template_path, engine='xlrd')
        else:
            template_df = pd.read_excel(template_path, engine='openpyxl')
        
        # 获取模板的列名（跳过前两行，因为前两行是说明）
        template_columns = list(template_df.columns)
        
        # 创建新的DataFrame，使用模板的列结构
        result_data = []
        
        # 字段映射关系
        field_mapping = {
            'name': 'name',
            'idtype': 'idtype',
            'id': 'id',
            'sex': 'sex',
            'birthdate': 'birthdate',
            'nationality': 'nationality',
            'polity': 'polity',
            'marital': 'marital',
            'joinworkdate': 'joinworkdate',
            'mobile': 'mobile',
            'usedname': 'usedname',
            'censusaddr': 'censusaddr',
            'permanreside': 'permanreside',
            'fileaddress': 'fileaddress',
            'joinpolitydate': 'joinpolitydate',
            'age': 'age',
            'workage': 'workage',
        }
        
        # 地址字段映射
        addr_mapping = {
            'addr.country': 'addr.country',
            'addr.province': 'addr.province',
            'addr.city': 'addr.city',
            'addr.county': 'addr.county',
            'addr.detail': 'addr.detail',
            'addr.postcode': 'addr.postcode',
        }
        
        for person_data in data_list:
            row = {}
            
            # 初始化所有字段为空
            for col in template_columns:
                row[col] = ''
            
            # 映射基本字段
            for json_key, template_key in field_mapping.items():
                if json_key in person_data and template_key in template_columns:
                    value = person_data[json_key]
                    if value:
                        row[template_key] = str(value).strip()
            
            # 映射地址字段
            if 'addr' in person_data:
                addr = person_data['addr']
                if isinstance(addr, dict):
                    for json_key, template_key in addr_mapping.items():
                        field_name = json_key.split('.')[1]  # 获取 'country', 'province' 等
                        if field_name in addr and template_key in template_columns:
                            value = addr[field_name]
                            if value:
                                row[template_key] = str(value).strip()
                elif isinstance(addr, str):
                    # 如果地址是字符串，尝试解析
                    # 可以进一步优化地址解析逻辑
                    if 'addr.detail' in template_columns:
                        row['addr.detail'] = addr
            
            # 处理嵌套的地址字段（如果JSON中直接有 addr.country 这样的键）
            for json_key, template_key in addr_mapping.items():
                if json_key in person_data and template_key in template_columns:
                    value = person_data[json_key]
                    if value:
                        row[template_key] = str(value).strip()
            
            result_data.append(row)
        
        # 创建DataFrame
        result_df = pd.DataFrame(result_data, columns=template_columns)
        
        # 复制模板的前两行（说明行）到结果中
        template_header = template_df.iloc[:2].copy()
        
        # 确保说明行的列数与模板列数一致
        if len(template_header) > 0:
            # 合并说明行和数据行
            final_df = pd.concat([template_header, result_df], ignore_index=True)
        else:
            final_df = result_df
        
        logger.info(f"成功映射 {len(result_data)} 条数据到模板格式")
        return final_df
        
    except Exception as e:
        logger.error(f"映射到模板格式失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def save_to_excel(df, output_path):
    """保存DataFrame到Excel"""
    try:
        df.to_excel(output_path, index=False, engine='openpyxl')
        print(f"✓ 表格已保存到: {output_path}")
        logger.info(f"表格已保存到: {output_path}, 共 {len(df)} 行, {len(df.columns)} 列")
        return True
    except Exception as e:
        logger.error(f"保存Excel失败: {str(e)}")
        raise


def save_to_template_format(df, output_path, template_path):
    """保存DataFrame到模板格式的Excel文件，完全保留模板的结构、表头和所有工作表"""
    try:
        import shutil
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        # 先复制模板文件
        print(f"正在复制模板文件...")
        shutil.copy2(template_path, output_path)
        
        # 使用openpyxl打开复制的文件
        wb = load_workbook(output_path)
        
        # 获取第一个工作表（通常是数据工作表）
        # 如果模板有多个工作表，保留所有工作表，只在第一个工作表写入数据
        if len(wb.sheetnames) > 0:
            ws = wb[wb.sheetnames[0]]
            print(f"正在写入数据到工作表: {ws.title}")
        else:
            raise Exception("模板文件没有工作表")
        
        # 确定数据开始行（跳过表头，通常是前2行）
        data_start_row = 3  # 第3行开始写入数据（第1、2行是表头）
        
        # 获取列名映射（df的列名对应模板的列位置）
        template_columns = list(df.columns)
        
        # 检查DataFrame是否包含表头行（前2行可能是表头说明）
        # 如果前2行的数据看起来像表头（包含"导入信息"、"校验信息"等），则跳过
        data_df = df.copy()
        if len(df) >= 2:
            # 检查第一行是否包含表头关键词
            first_row_values = [str(v).strip() for v in df.iloc[0].values if pd.notna(v)]
            if any(keyword in ' '.join(first_row_values) for keyword in ['导入信息', '校验信息', '导入', '校验']):
                # 跳过前2行（表头）
                data_df = df.iloc[2:].copy()
                print(f"检测到表头行，将从第3行开始写入数据")
        
        # 清除原有数据（从第3行开始清除，保留表头）
        max_row = ws.max_row
        if max_row >= data_start_row:
            ws.delete_rows(data_start_row, max_row - data_start_row + 1)
        
        # 写入新数据
        for row_idx, (_, row_data) in enumerate(data_df.iterrows(), start=data_start_row):
            for col_idx, col_name in enumerate(template_columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data[col_name]
                # 处理NaN值
                if pd.isna(value):
                    cell.value = ''
                else:
                    cell.value = str(value) if value is not None else ''
        
        # 保存文件
        wb.save(output_path)
        wb.close()
        
        print(f"✓ 已保存到模板格式: {output_path}")
        print(f"  保留了 {len(wb.sheetnames)} 个工作表")
        logger.info(f"已保存到模板格式: {output_path}, 共 {len(df)} 行数据, {len(wb.sheetnames)} 个工作表")
        return True
    except Exception as e:
        logger.error(f"保存模板格式Excel失败: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def load_json_data(json_path):
    """从JSON文件加载数据"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 确保返回的是列表格式
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            return [data]
        else:
            raise ValueError("JSON数据格式不正确，应为数组或对象")
    except Exception as e:
        logger.error(f"加载JSON文件失败: {str(e)}")
        raise


def json_to_template(json_path, template_path=None, output_path=None):
    """从JSON文件读取数据并写入到模板中"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 加载JSON数据
        print(f"正在从JSON文件加载数据: {os.path.basename(json_path)}")
        data_list = load_json_data(json_path)
        print(f"✓ 成功加载 {len(data_list)} 条数据")
        
        # 查找模板文件
        if template_path is None:
            template_path = None
            possible_template_names = [
                "首图职称参评信息导入模版.xls",

            ]
            
            # 先尝试精确匹配
            for name in possible_template_names:
                path = os.path.join(script_dir, name)
                if os.path.exists(path):
                    template_path = path
                    break
            
            # 如果精确匹配失败，尝试模糊匹配
            if not template_path:
                template_files = list(Path(script_dir).glob("*导入模版*"))
                if template_files:
                    template_path = str(template_files[0])
        
        if not template_path or not os.path.exists(template_path):
            raise FileNotFoundError(f"未找到模板文件，请确保模板文件存在于 {script_dir} 目录")
        
        print(f"正在使用模板文件: {os.path.basename(template_path)}")
        
        # 映射数据到模板格式
        print("正在将数据映射到模板格式...")
        df = map_to_template_format(data_list, template_path)
        
        # 确定输出文件名
        if output_path is None:
            # 从数据中提取姓名
            name_for_filename = "人员信息"
            if len(data_list) > 0:
                first_person = data_list[0]
                if isinstance(first_person, dict) and 'name' in first_person:
                    name_for_filename = str(first_person['name']).strip()
                    name_for_filename = re.sub(r'[<>:"/\\|?*]', '', name_for_filename)
                    if not name_for_filename:
                        name_for_filename = "人员信息"
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(script_dir, f"{name_for_filename}_{timestamp}.xlsx")
        
        # 保存到模板格式（传入模板路径以完全复制模板）
        save_to_template_format(df, output_path, template_path)
        
        print(f"\n✓ 处理完成！")
        print(f"  输入JSON: {os.path.basename(json_path)}")
        print(f"  模板文件: {os.path.basename(template_path)}")
        print(f"  输出文件: {os.path.basename(output_path)}")
        print(f"  数据条数: {len(data_list)}")
        
        return output_path
        
    except Exception as e:
        error_msg = f"从JSON写入模板失败: {str(e)}"
        print(f"错误: {error_msg}")
        logger.error(error_msg)
        import traceback
        logger.error(traceback.format_exc())
        raise


def main():
    """主函数"""
    driver = None
    
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 查找要上传的文件（PDF、图片等）
        file_extensions = ['.pdf', '.png', '.jpg', '.jpeg', '.txt', '.doc', '.docx']
        files = []
        for ext in file_extensions:
            files.extend(list(Path(script_dir).glob(f"*{ext}")))
        
        if not files:
            print(f"错误: 在 {script_dir} 目录下未找到可上传的文件")
            print(f"支持的文件格式: {', '.join(file_extensions)}")
            return
        
        # 自动选择第一个文件（优先选择PDF）
        pdf_files = [f for f in files if f.suffix.lower() == '.pdf']
        if pdf_files:
            file_path = pdf_files[0]
            print(f"\n找到 {len(files)} 个文件，自动选择PDF文件: {file_path.name}")
        else:
            file_path = files[0]
            print(f"\n找到 {len(files)} 个文件，自动选择: {file_path.name}")
        
        print(f"\n正在处理文件: {file_path.name}")
        logger.info(f"开始处理文件: {file_path}")
        
        # 启动浏览器
        print("正在启动浏览器...")
        driver = setup_driver()
        
        # 打开豆包聊天页面（无需登录）
        print("正在打开豆包聊天页面...")
        driver.get("https://www.doubao.com/")
        time.sleep(3)
        
        # 先查找模板文件
        template_path = None
        possible_template_names = [
            "首图职称参评信息导入模版.xls",
        ]
        
        # 先尝试精确匹配
        for name in possible_template_names:
            path = os.path.join(script_dir, name)
            if os.path.exists(path):
                template_path = path
                break
        
        # 如果精确匹配失败，尝试模糊匹配（包含"导入模版"的文件）
        if not template_path:
            template_files = list(Path(script_dir).glob("*导入模版*"))
            if template_files:
                template_path = str(template_files[0])
                print(f"找到模板文件: {os.path.basename(template_path)}")
        
        # 上传PDF文件
        upload_file_to_doubao(driver, str(file_path), wait_after_upload=True)
        
        # 如果找到模板文件，也上传模板
        if template_path:
            print(f"正在上传模板文件: {os.path.basename(template_path)}")
            time.sleep(1)  # 等待一下，确保上一个文件上传完成
            upload_file_to_doubao(driver, template_path, wait_after_upload=True)
        
        # 发送消息（先只提取内容，不整理）
        message = "把pdf的内容整理到导入模板里"
        send_message_to_doubao(driver, message)
        
        # 等待响应
        wait_for_response(driver, timeout=120)
        
        # 提取结构化数据
        data, data_type = extract_structured_data_from_response(driver)
        
        if data is not None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 从数据中提取姓名用于文件名
            name_for_filename = "人员信息"  # 默认名称
            if data_type == 'json' and len(data) > 0:
                # 尝试从第一条数据中获取姓名
                first_person = data[0] if isinstance(data, list) else data
                if isinstance(first_person, dict) and 'name' in first_person:
                    name_for_filename = str(first_person['name']).strip()
                    # 清理文件名中的非法字符
                    name_for_filename = re.sub(r'[<>:"/\\|?*]', '', name_for_filename)
                    if not name_for_filename:
                        name_for_filename = "人员信息"
                
                # 保存JSON数据到文件（先只保存提取的内容，不整理到模板）
                json_output_path = os.path.join(script_dir, f"{name_for_filename}_{timestamp}.json")
                try:
                    with open(json_output_path, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    print(f"\n✓ 提取的内容已保存到: {os.path.basename(json_output_path)}")
                    logger.info(f"JSON数据已保存到: {json_output_path}")
                    
                    # 显示提取的内容摘要
                    print(f"\n提取的内容预览：")
                    print(f"  共提取了 {len(data)} 条人员信息")
                    if len(data) > 0:
                        print(f"\n第一条信息示例：")
                        first_person = data[0]
                        for key, value in first_person.items():
                            if value:  # 只显示有值的字段
                                print(f"    {key}: {value}")
                    
                    print(f"\n请检查 {os.path.basename(json_output_path)} 文件，确认提取的内容是否正确。")
                    print(f"如果内容正确，可以运行以下命令整理到模板：")
                    print(f"  python send_to_doubao_extract_table.py {json_output_path}")
                    
                except Exception as e:
                    logger.warning(f"保存JSON文件失败: {str(e)}")
            elif data_type == 'table' and len(data) > 0:
                # 尝试从表格中提取姓名
                # 方法1: 查找"姓名"列
                name_col = None
                for col in data.columns:
                    if '姓名' in str(col) or 'name' in str(col).lower():
                        name_col = col
                        break
                
                if name_col and len(data) > 0:
                    first_name = data.iloc[0][name_col] if pd.notna(data.iloc[0][name_col]) else None
                    if first_name and str(first_name).strip():
                        name_for_filename = str(first_name).strip()
                        name_for_filename = re.sub(r'[<>:"/\\|?*]', '', name_for_filename)
                
                # 方法2: 如果没找到姓名列，尝试从第一行第一列获取
                if not name_for_filename or name_for_filename == "人员信息":
                    if len(data) > 0 and len(data.columns) > 0:
                        first_cell = data.iloc[0, 0]
                        if pd.notna(first_cell) and str(first_cell).strip():
                            potential_name = str(first_cell).strip()
                            # 检查是否像姓名（2-4个中文字符）
                            if re.match(r'^[\u4e00-\u9fa5]{2,4}$', potential_name):
                                name_for_filename = potential_name
                                name_for_filename = re.sub(r'[<>:"/\\|?*]', '', name_for_filename)
                
                # 如果找到模板，将表格数据写入模板
                if template_path and os.path.exists(template_path):
                    print(f"\n正在将表格数据写入模板...")
                    output_path = os.path.join(script_dir, f"{name_for_filename}_{timestamp}.xlsx")
                    
                    # 复制模板并写入数据
                    try:
                        import shutil
                        from openpyxl import load_workbook
                        
                        # 复制模板文件
                        print(f"正在复制模板文件...")
                        shutil.copy2(template_path, output_path)
                        
                        # 打开复制的文件
                        wb = load_workbook(output_path)
                        ws = wb[wb.sheetnames[0]]  # 使用第一个工作表
                        
                        # 确定数据开始行（跳过表头，通常是前2行）
                        data_start_row = 3
                        
                        # 清除原有数据（从第3行开始）
                        max_row = ws.max_row
                        if max_row >= data_start_row:
                            ws.delete_rows(data_start_row, max_row - data_start_row + 1)
                        
                        # 写入表格数据
                        # 获取模板的列名（从第2行读取，因为第1行是说明）
                        template_headers = []
                        if ws.max_row >= 2:
                            for col_idx in range(1, ws.max_column + 1):
                                cell_value = ws.cell(row=2, column=col_idx).value
                                if cell_value:
                                    template_headers.append(str(cell_value).strip())
                        
                        # 将DataFrame的列名映射到模板列
                        df_columns = list(data.columns)
                        
                        # 写入数据
                        for row_idx, (_, row_data) in enumerate(data.iterrows(), start=data_start_row):
                            for col_idx, template_col in enumerate(template_headers, start=1):
                                # 尝试匹配列名
                                value = ''
                                for df_col in df_columns:
                                    # 模糊匹配列名
                                    if template_col in str(df_col) or str(df_col) in template_col:
                                        cell_value = row_data[df_col]
                                        if pd.notna(cell_value):
                                            value = str(cell_value).strip()
                                        break
                                
                                # 如果没找到匹配，尝试按位置匹配
                                if not value and col_idx - 1 < len(df_columns):
                                    cell_value = row_data.iloc[col_idx - 1]
                                    if pd.notna(cell_value):
                                        value = str(cell_value).strip()
                                
                                # 写入单元格
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.value = value
                        
                        # 保存文件
                        wb.save(output_path)
                        wb.close()
                        
                        print(f"✓ 已保存到模板格式: {os.path.basename(output_path)}")
                        print(f"  保留了 {len(wb.sheetnames)} 个工作表")
                        print(f"  写入了 {len(data)} 行数据")
                        logger.info(f"表格数据已写入模板: {output_path}, 共 {len(data)} 行")
                        
                    except Exception as e:
                        logger.error(f"写入模板失败: {str(e)}")
                        import traceback
                        logger.error(traceback.format_exc())
                        # 如果写入模板失败，至少保存原始表格
                        table_output_path = os.path.join(script_dir, f"{name_for_filename}_{timestamp}_原始表格.xlsx")
                        save_to_excel(data, table_output_path)
                        print(f"⚠️  写入模板失败，已保存原始表格到: {os.path.basename(table_output_path)}")
        else:
                    # 如果没有模板，直接保存表格
                    table_output_path = os.path.join(script_dir, f"{name_for_filename}_{timestamp}.xlsx")
                    save_to_excel(data, table_output_path)
                    print(f"\n✓ 提取的表格已保存到: {os.path.basename(table_output_path)}")
                    print(f"  共 {len(data)} 行 × {len(data.columns)} 列")
                    print(f"⚠️  未找到模板文件，已保存为普通Excel格式")
            
            return  # 处理完成
        else:
            print("\n⚠️  未能提取到数据")
            print("提示: 请确保豆包返回了结构化数据（JSON或表格格式）")
        
        # 保持浏览器打开
        print("\n浏览器将保持打开10秒，您可以查看结果...")
        time.sleep(10)
        
    except Exception as e:
        error_msg = f"处理失败: {str(e)}"
        print(f"错误: {error_msg}")
        logger.error(error_msg)
        import traceback
        logger.error(traceback.format_exc())
        
    finally:
        # 关闭浏览器
        if driver:
            try:
                driver.quit()
            except:
                pass


if __name__ == "__main__":
    import sys
    
    # 如果提供了JSON文件路径作为参数，直接从JSON写入模板
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
        if os.path.exists(json_path):
            try:
                json_to_template(json_path)
            except Exception as e:
                print(f"错误: {str(e)}")
                sys.exit(1)
        else:
            print(f"错误: JSON文件不存在: {json_path}")
            sys.exit(1)
    else:
        # 默认行为：从豆包提取数据
    main()

